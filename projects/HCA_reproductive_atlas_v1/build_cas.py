#!/usr/bin/env python3
"""Draft CAS builder for HCA Female Reproductive System Cell Atlas v1.

Decisions (agreed in discussion):
- Labelsets = L1..L4 nomenclature tiers from cell_ontology_mapping.xlsx (the 1:1,
  disambiguated source). rank: L1=3, L2=2, L3=1, L4=0.
- Leaf identity = object celltype_HCA_fine; L4 is (essentially) its fullname.
- Skip-level parents, ragged tree, NO padding, NO level 5, materialise every tier.
- Source-dataset cell-type labels -> transferred_annotations (cell_count/cell_ratio).
- Everything else the object owns (celltype_HCA* codes + tissue/stage/disease
  descriptors) -> author_annotations, each a {value, cell_count, cell_ratio}
  distribution (ratio 1 where homogeneous, e.g. the defining fine code at a leaf).
- CL term (from the sheet) attaches at each code's terminal node.

Throwaway/draft. Reads obs.categoricals.parquet + the two xlsx; writes cas_draft.json.
"""
import openpyxl, pandas as pd, json, re
from collections import defaultdict

NULLS = {None, "", "nan", "None"}
def cln(x):
    if x is None: return None
    s = str(x).strip()
    return None if s.lower() in {"", "nan", "none"} else s

# ---------------------------------------------------------------- master L1-L4
wb = openpyxl.load_workbook("inputs/cell_ontology_mapping.xlsx"); ws = wb["Final"]
# Canonicalise trivial nomenclature-variant that forks the tree. AUDIT (all L1-L4) found
# exactly ONE: the fetal sub-spreadsheet spells L2 'Stromal-interstitial fibroblast' (singular)
# while the postnatal one uses 'fibroblasts' (plural) — this split the stromal lineage into a
# fetal tree and an adult tree and duplicated 4 L3 region nodes. Merge singular -> plural so
# fetal + adult stroma unify under one L2 (and the duplicated L3s / artefactual chains dissolve).
CANON = {"Stromal-interstitial fibroblast": "Stromal-interstitial fibroblasts"}
sheet = {}
for r in list(ws.iter_rows(values_only=True))[1:]:
    leaf = cln(r[0])
    if not leaf: continue
    path = [cln(r[2]), cln(r[3]), cln(r[4]), cln(r[5])]        # L1,L2,L3,L4
    path = [CANON.get(v, v) for v in path]
    sheet[leaf] = {"path": path, "cl_id": cln(r[10]), "cl_label": cln(r[6]) or cln(r[8])}
wb.close()

# ---------------------------------------------------------------- object obs
obs = pd.read_parquet("h5ad_obs/obs.categoricals.parquet").astype(str)
N_TOTAL = len(obs)
objcodes = set(obs["celltype_HCA_fine"].unique())

# join object code -> sheet row, with light normalisation for known drifts
def resolve(code):
    if code in sheet: return code
    if code.rstrip("+") in sheet: return code.rstrip("+")          # Endo_cap_APCDD1+
    for a, b in (("Fibs", "Fib"), ("Fib", "Fibs")):               # singular/plural drift
        if code.replace(a, b) in sheet: return code.replace(a, b)
    return None

code2key = {c: resolve(c) for c in objcodes}
unmatched = sorted(c for c, k in code2key.items() if k is None)
print(f"object fine codes: {len(objcodes)} | matched to sheet: {sum(v is not None for v in code2key.values())} | unmatched: {unmatched}")

# per-code path (list of 4, trailing None allowed)
code2path = {}
for c in objcodes:
    k = code2key[c]
    code2path[c] = sheet[k]["path"] if k else [None, None, None, None]

# ---------------------------------------------------------------- build nodes
# node key = tuple(path[:level]); level = len(tuple); parent = tuple(path[:level-1])
def node_keys_for_path(path):
    keys = []
    for i in range(4):
        if path[i] is None:  # ragged: stop at first blank (blanks only at deep end)
            break
        keys.append(tuple(path[: i + 1]))
    return keys

nodes = {}   # key -> {level,label,parent,codes_terminal}
for c, path in code2path.items():
    ks = node_keys_for_path(path)
    for i, k in enumerate(ks):
        if k not in nodes:
            nodes[k] = {"level": i + 1, "label": k[-1],
                        "parent": k[:-1] if i > 0 else None, "codes_terminal": []}
    if ks:                        # this code terminates at its deepest node
        nodes[ks[-1]]["codes_terminal"].append(c)

# accession registry
key2acc = {k: f"HCArepro:L{v['level']}:{i:04d}" for i, (k, v) in enumerate(nodes.items())}

# CL term per node (from the code(s) that terminate there)
for k, v in nodes.items():
    cls = {(sheet[code2key[c]]["cl_id"], sheet[code2key[c]]["cl_label"])
           for c in v["codes_terminal"] if code2key[c]}
    v["cl"] = list(cls)

# ---------------------------------------------------------------- assign cells to nodes per level
# vectorised: precompute per-code cumulative node-key strings, then .map the single
# celltype_HCA_fine column (fast) instead of a row-wise apply over 2.24M rows.
code2nodestr = {c: {} for c in objcodes}   # c -> {level: "a|b|c" or None}
for c, path in code2path.items():
    for i in range(1, 5):
        pref = path[:i]
        code2nodestr[c][i] = None if any(p is None for p in pref) else "|".join(pref)
for i in range(1, 5):
    obs[f"__node{i}"] = obs["celltype_HCA_fine"].map({c: code2nodestr[c][i] for c in objcodes})

# ---------------------------------------------------------------- field groups
TRANSFER = ["celltype_GarciaAlonso2021","celltype_GarciaAlonso2022","celltype_Ulrich2022",
            "celltype_Ulrich2024","celltype_Weigert2025","celltype_OvarySanger2026",
            "celltype_Lardenois2026","celltype_Lorenzi2025","celltype_HECA"]
TRANSFER_DOI = {
 "celltype_GarciaAlonso2021":"10.1038/s41588-021-00972-2",
 "celltype_GarciaAlonso2022":"10.1038/s41586-022-04918-4",
 "celltype_Ulrich2022":"10.1016/j.devcel.2022.02.017",
 "celltype_Ulrich2024":"10.1073/pnas.2404775121",
 "celltype_Weigert2025":"10.1038/s41467-024-55440-2",
 "celltype_OvarySanger2026":None,   # newly-generated Sanger ovary data (no external DOI)
 "celltype_Lardenois2026":"10.1016/j.devcel.2025.09.011",
 "celltype_Lorenzi2025":"10.1038/s41586-025-09875-2",
 "celltype_HECA":"10.1038/s41588-024-01873-w",   # tentative
}
AUTHOR_CELLTYPE = ["celltype_HCA_fine","celltype_HCA","celltype_HCA_broad","celltype_HCA_lineage"]
AUTHOR_DESCRIPTOR = ["Organ","Organ_part","Tissue_ROI","Target_cell_population",
    "Developmental_stage","Postnatal_age_years","Tanner Stage","Gestational_age_pcw","Menstrual_stage",
    "Disease","Clinical_diagnosis","Observed_pathology","Sampled_site_condition"]

# ---------------------------------------------------------------- per-node distributions
def dist_by_node(level_col, field, drop_unknown):
    """return {node_strkey: [ {value, cell_count} ... ]}"""
    sub = obs[[level_col, field]].dropna(subset=[level_col])
    sub = sub[sub[level_col] != "None"]
    g = sub.groupby([level_col, field], observed=True).size()
    out = defaultdict(list)
    for (nk, val), cnt in g.items():
        if val in NULLS or val == "nan": continue
        if drop_unknown and val.lower() in {"unknown","not applicable"}: continue
        out[nk].append((val, int(cnt)))
    return out

ncells = {}
for i in range(1, 5):
    col = f"__node{i}"
    vc = obs[col].dropna()
    vc = vc[vc != "None"].value_counts()
    for nk, n in vc.items(): ncells[nk] = int(n)

# precompute distributions keyed (field -> {nodestr -> [(val,cnt)]})
author_dist = {f: {} for f in AUTHOR_CELLTYPE + AUTHOR_DESCRIPTOR}
transfer_dist = {f: {} for f in TRANSFER}
for i in range(1, 5):
    col = f"__node{i}"
    for f in AUTHOR_CELLTYPE + AUTHOR_DESCRIPTOR:
        d = dist_by_node(col, f, drop_unknown=False)
        for nk, lst in d.items(): author_dist[f][nk] = lst
    for f in TRANSFER:
        d = dist_by_node(col, f, drop_unknown=True)
        for nk, lst in d.items(): transfer_dist[f][nk] = lst

# ---------------------------------------------------------------- assemble annotations
def acc_of(key): return key2acc[key]
RANK = {1:3, 2:2, 3:1, 4:0}

annotations = []
for key, v in nodes.items():
    nk = "|".join([x for x in key])
    n = ncells.get(nk, 0)
    lvl = v["level"]
    ann = {
        "labelset": f"L{lvl}",
        "rank": RANK[lvl],
        "cell_label": v["label"],
        "cell_set_accession": acc_of(key),
        "parent_cell_set_accession": acc_of(v["parent"]) if v["parent"] else None,
        "n_cells": n,
    }
    # CL term at terminal nodes
    if v["cl"]:
        ids = [c for c in v["cl"] if c[0]]
        if len(ids) == 1:
            ann["cell_ontology_term_id"] = ids[0][0]
            ann["cell_ontology_term"]   = ids[0][1]
        elif len(ids) > 1:
            ann["cell_ontology_term_id"] = [c[0] for c in ids]   # ambiguous -> list (flagged)
    # author annotations (own labels + descriptors), each a value distribution
    aa = []
    for f in AUTHOR_CELLTYPE + AUTHOR_DESCRIPTOR:
        lst = author_dist[f].get(nk, [])
        if not lst: continue
        vals = [{"value": val, "cell_count": cnt, "cell_ratio": round(cnt / n, 4)}
                for val, cnt in sorted(lst, key=lambda t: -t[1])]
        aa.append({"field": f, "values": vals})
    ann["author_annotations"] = aa
    # transferred annotations (other-source cell-type labels only)
    ta = []
    for f in TRANSFER:
        lst = transfer_dist[f].get(nk, [])
        for val, cnt in sorted(lst, key=lambda t: -t[1]):
            ta.append({"transferred_cell_label": val, "source_labelset": f,
                       "source_taxonomy": (f"DOI:{TRANSFER_DOI[f]}" if TRANSFER_DOI[f] else "Sanger (newly generated; no DOI)"),
                       "cell_count": cnt, "cell_ratio": round(cnt / n, 4)})
    if ta: ann["transferred_annotations"] = ta
    annotations.append(ann)

# ---------------------------------------------------------------- mint generic leaves for mixed nodes
# A "mixed" node is an internal (parent) node that also has object codes terminating on it
# (a generic/un-subtyped cell type coexisting with finer subtypes). Each such object code is a
# real celltype_HCA_fine leaf, so give it its own leaf node; the parent becomes a pure supertype.
def dist_for_code(code, fields, drop_unknown):
    sub = obs[obs["celltype_HCA_fine"].astype(str) == code]
    n = len(sub); res = []
    for f in fields:
        vc = sub[f].astype(str).value_counts()
        vals = [{"value": v, "cell_count": int(c), "cell_ratio": round(c / n, 4)}
                for v, c in vc.items()
                if v not in NULLS and v != "nan"
                and not (drop_unknown and v.lower() in {"unknown", "not applicable"})]
        if vals: res.append((f, vals))
    return n, res

by_acc = {a["cell_set_accession"]: a for a in annotations}
child_parent = set(v["parent"] for v in nodes.values() if v["parent"])
minted = []
mint_i = 0
for key, v in nodes.items():
    if not (key in child_parent and v["codes_terminal"]):
        continue
    parent_acc = key2acc[key]
    parent_name = v["label"]
    parent_lvl = v["level"]
    for code in v["codes_terminal"]:
        n, aa_raw = dist_for_code(code, AUTHOR_CELLTYPE + AUTHOR_DESCRIPTOR, drop_unknown=False)
        _, ta_raw = dist_for_code(code, TRANSFER, drop_unknown=True)
        k = code2key.get(code)
        cl_id = sheet[k]["cl_id"] if k else None
        cl_lab = sheet[k]["cl_label"] if k else None
        leaf_lvl = parent_lvl + 1
        acc = f"HCArepro:L{leaf_lvl}:m{mint_i:03d}"; mint_i += 1
        leaf = {
            "labelset": f"L{leaf_lvl}",
            "rank": RANK.get(leaf_lvl, 0),
            "cell_label": parent_name,   # an un-subtyped X is still an X; fullname inherited from supertype
            "cell_set_accession": acc,
            "parent_cell_set_accession": parent_acc,
            "n_cells": n,
            "comment": (f"Minted generic leaf for celltype_HCA_fine='{code}' ({n} cells). The master "
                        f"L1-L4 nomenclature terminates at L{parent_lvl} '{parent_name}' with no finer "
                        f"label, so these author-labelled cells — pericytes/etc. the annotators did NOT "
                        f"assign to a finer subtype — had no leaf of their own; the parent node is retained "
                        f"as a pure supertype. cell_label inherited from the supertype nomenclature."),
        }
        if cl_id:
            leaf["cell_ontology_term_id"] = cl_id
            leaf["cell_ontology_term"] = cl_lab
        leaf["author_annotations"] = [
            {"field": f, "values": vals} for f, vals in aa_raw]
        ta = []
        for f, vals in ta_raw:
            for vv in vals:
                ta.append({"transferred_cell_label": vv["value"], "source_labelset": f,
                           "source_taxonomy": (f"DOI:{TRANSFER_DOI[f]}" if TRANSFER_DOI[f] else "Sanger (newly generated; no DOI)"),
                           "cell_count": vv["cell_count"], "cell_ratio": vv["cell_ratio"]})
        if ta: leaf["transferred_annotations"] = ta
        minted.append(leaf)
        # annotate the parent as now-pure supertype
        by_acc[parent_acc]["comment"] = (
            f"Supertype/internal node. Its un-subtyped members were split into minted generic "
            f"leaf/leaves (see children with 'Minted generic leaf' comments). n_cells is the full subtree.")

annotations.extend(minted)

# ---------------------------------------------------------------- enrich leaves with media-4 fields
# Markers_positive -> marker_gene_evidence; Markers_negative -> negative_marker_gene_evidence;
# Alternative_celltype_labels -> synonyms; Celltype_description -> comment (mixed bag; mine later);
# author-year citations extracted from the description -> rationale_citations (raw) for DOI resolution.
import re as _re
wb = openpyxl.load_workbook("inputs/media-4.xlsx"); ws = wb["B. HCA_femaleRepSys_v1_celltype"]
m4 = {}
for r in list(ws.iter_rows(values_only=True))[6:]:
    code = cln(r[0])
    if not code: continue
    split = lambda s: [g.strip() for g in _re.split(r"[;,]", str(s)) if g and g.strip()]
    m4[code] = {"pos": split(r[5]) if r[5] else [], "neg": split(r[6]) if r[6] else [],
                "desc": cln(r[7]), "syn": split(r[8]) if r[8] else [],
                "path": [CANON.get(v, v) for v in [cln(r[1]), cln(r[2]), cln(r[3]), cln(r[4])]]}
wb.close()
CITE = _re.compile(r"([A-Z][A-Za-z\-]+(?:\s+(?:and|&)\s+[A-Z][A-Za-z\-]+)?\s+et al\.?,?\s*(?:19|20)\d{2})")
# defining code -> annotation (ratio-1 celltype_HCA_fine == leaves + minted leaves)
# also index by a normalised key (Fibs->Fib, strip trailing '+') to absorb known label drift
# between media-4 (singular 'Fib') and the object (plural 'Fibs').
def _norm(s): return s.replace("Fibs", "Fib").rstrip("+")
code2ann = {}; norm2ann = {}
for a in annotations:
    fa = [x for x in a.get("author_annotations", []) if x["field"] == "celltype_HCA_fine"]
    if fa and len(fa[0]["values"]) == 1 and fa[0]["values"][0]["cell_ratio"] == 1.0:
        v = fa[0]["values"][0]["value"]
        code2ann[v] = a
        norm2ann.setdefault(_norm(v), a)
attached = 0; unattached = []
for code, m in m4.items():
    a = code2ann.get(code) or norm2ann.get(_norm(code))
    if a is None:
        if m["pos"] or m["desc"]: unattached.append(code)
        continue
    if m["pos"]: a["marker_gene_evidence"] = m["pos"]
    if m["neg"]: a["negative_marker_gene_evidence"] = m["neg"]
    if m["syn"]: a["synonyms"] = m["syn"]
    if m["desc"]:
        a["comment"] = ((a.get("comment", "") + " | ") if a.get("comment") else "") + "media-4 description: " + m["desc"]
        cites = CITE.findall(m["desc"])
        if cites: a["rationale_citations"] = cites
    attached += 1

# second pass: broad/supertype media-4 codes -> internal nomenclature nodes, matched by L-path terminal
node_idx = defaultdict(list)
for a in annotations:
    node_idx[(a["labelset"], a["cell_label"])].append(a)
def is_ratio1_leaf(a):
    fa = [x for x in a.get("author_annotations", []) if x["field"] == "celltype_HCA_fine"]
    return bool(fa) and len(fa[0]["values"]) == 1 and fa[0]["values"][0]["cell_ratio"] == 1.0
broad_attached = 0; still_unattached = []
for code in unattached:
    m = m4[code]; path = m["path"]
    depth = max([i for i, v in enumerate(path, 1) if v] or [0])
    term = path[depth - 1] if depth else None
    cand = [x for x in node_idx.get((f"L{depth}", term), []) if not is_ratio1_leaf(x)]
    if len(cand) == 1:
        a = cand[0]
        if m["pos"] and not a.get("marker_gene_evidence"): a["marker_gene_evidence"] = m["pos"]
        if m["neg"] and not a.get("negative_marker_gene_evidence"): a["negative_marker_gene_evidence"] = m["neg"]
        if m["syn"] and not a.get("synonyms"): a["synonyms"] = m["syn"]
        if m["desc"]:
            a["comment"] = ((a.get("comment", "") + " | ") if a.get("comment") else "") + "media-4 (supertype) description: " + m["desc"]
            cites = CITE.findall(m["desc"])
            if cites: a["rationale_citations"] = a.get("rationale_citations", []) + cites
        a["marker_source"] = f"media-4 broad code '{code}' (supertype-level)"
        broad_attached += 1
    else:
        still_unattached.append(code)
print(f"media-4 fields attached to {attached} leaves | broad->internal: {broad_attached} | "
      f"still unattached ({len(still_unattached)}): {sorted(still_unattached)}")
n_cites = sum(1 for a in annotations if a.get("rationale_citations"))
print(f"annotations with extracted citations: {n_cites}")

# ---- resolve mined citations -> rationale_dois (DOIs VERIFIED via literature search, not from memory)
CITE_DOI = {   # high-confidence matches only
    "Villani et al., 2017":    "10.1126/science.aah4573",       # blood DC/monocyte atlas; AS DC (AXL+SIGLEC6+)
    "Maier et al., 2020":      "10.1038/s41586-020-2134-y",     # mregDC regulatory program
    "Vento-Tormo et al., 2018":"10.1038/s41586-018-0698-6",     # maternal-fetal interface; uNK1/2/3
    "Nguyen et al 2017":       "10.1093/humrep/dex289",         # N-cadherin endometrial epithelial progenitors
    "Fu et al.":               "10.1038/s41586-025-08982-4",    # PRDM16+ tolDCs, gut antigen tolerance
}
CITE_UNRESOLVED = {"Wang et al., 2021", "Masopust et al., 2026"}   # no verified DOI (see build notes)
for a in annotations:
    desc = a.get("comment", "")
    dois = []
    for k, d in CITE_DOI.items():
        if k in desc and d not in dois: dois.append(d)
    if dois: a["rationale_dois"] = dois
    unres = [c for c in a.get("rationale_citations", []) if c in CITE_UNRESOLVED]
    if unres:
        a["comment"] = desc + f" | UNRESOLVED citation(s) {unres}: no verified DOI found."
n_doi = sum(1 for a in annotations if a.get("rationale_dois"))
print(f"annotations with resolved rationale_dois: {n_doi}")

# provenance comment on the one duplicate-label artifact node (Endo_cap_APCDD1 / Endo_cap_APCDD1+)
for a in annotations:
    fine = [x for x in a.get("author_annotations", []) if x["field"] == "celltype_HCA_fine"]
    if fine and {v["value"] for v in fine[0]["values"]} >= {"Endo_cap_APCDD1", "Endo_cap_APCDD1+"}:
        a["comment"] = ("Object data-quality artifact: the h5ad carries two celltype_HCA_fine labels "
                        "for the same cell type — 'Endo_cap_APCDD1' and 'Endo_cap_APCDD1+' (trailing '+'). "
                        "Merged here to one leaf via label normalisation; flag to authors.")

annotations.sort(key=lambda a: (a["rank"], a["cell_label"]))

# ---------------------------------------------------------------- document
doc = {
    "title": "Human Female Reproductive System Cell Atlas v1",
    "description": "Draft CAS(+) export. Labelsets = master L1-L4 nomenclature (cell_ontology_mapping.xlsx). "
                   "Source-dataset labels in transferred_annotations; object's own labels + descriptors in author_annotations. "
                   "Local extensions: per-annotation 'comment' (free-text provenance, agent-readable), and generic leaves "
                   "minted for un-subtyped celltype_HCA_fine codes that share a nomenclature node with finer subtypes.",
    "matrix_file_id": "https://cellgeni-share.cog.sanger.ac.uk/REQ-69024/integrated_scvi_all_tissues_cellxgene_filtered.h5ad",
    "cellannotation_schema_version": "0.1.0-local",
    "source": {
        "doi": "10.64898/2026.06.10.731198",
        "title": "An integrated multimodal pan-organ atlas of the female reproductive system across the lifespan contextualises gynaecological pathologies",
        "authors": ["Celeste E Cohen", "Luz Garcia-Alonso", "Roser Vento-Tormo"],
    },
    "labelsets": [
        {"name": "L4", "rank": 0, "description": "Master nomenclature level 4 (finest; ≈ fullname of celltype_HCA_fine)"},
        {"name": "L3", "rank": 1, "description": "Master nomenclature level 3"},
        {"name": "L2", "rank": 2, "description": "Master nomenclature level 2"},
        {"name": "L1", "rank": 3, "description": "Master nomenclature level 1 (lineage)"},
    ],
    "annotations": annotations,
}

with open("cas_draft.json", "w") as f:
    json.dump(doc, f, indent=2, ensure_ascii=False)

# ---------------------------------------------------------------- validation summary
import os
print(f"\nwrote cas_draft.json ({os.path.getsize('cas_draft.json')/1e6:.1f} MB)")
by_rank = defaultdict(int)
for a in annotations: by_rank[a["labelset"]] += 1
print("annotations per labelset:", dict(by_rank), "| total nodes:", len(annotations))
n_minted = sum(1 for a in annotations if "m" in a["cell_set_accession"].split(":")[-1])
n_comment = sum(1 for a in annotations if a.get("comment"))
print(f"minted generic leaves: {n_minted} | annotations with a comment: {n_comment}")
# n_cells conservation: sum of L1 nodes == total object cells (every cell has L1)
l1sum = sum(a["n_cells"] for a in annotations if a["labelset"]=="L1")
print(f"Σ n_cells at L1 = {l1sum:,} (object = {N_TOTAL:,})  {'OK' if l1sum==N_TOTAL else 'MISMATCH'}")
# terminal nodes with a CL term
withcl = sum(1 for a in annotations if "cell_ontology_term_id" in a)
ambig = [a["cell_label"] for a in annotations if isinstance(a.get("cell_ontology_term_id"), list)]
print(f"nodes with CL term: {withcl} | ambiguous (multi-CL terminal): {ambig}")
# ratio-1 defining fine codes at leaves (sanity: L4 nodes where celltype_HCA_fine is single value ratio 1)
leaf_ratio1 = 0
for a in annotations:
    if a["labelset"]!="L4": continue
    fine=[x for x in a["author_annotations"] if x["field"]=="celltype_HCA_fine"]
    if fine and len(fine[0]["values"])==1 and fine[0]["values"][0]["cell_ratio"]==1.0: leaf_ratio1+=1
print(f"L4 nodes with a single celltype_HCA_fine @ratio 1.0: {leaf_ratio1}")
# nodes that are BOTH terminal-for-a-code and internal (mixed) -> flag
mixed=[]
child_parent=set(v["parent"] for v in nodes.values() if v["parent"])
for k,v in nodes.items():
    if v["codes_terminal"] and k in child_parent: mixed.append((v["label"], v["level"], v["codes_terminal"]))
print(f"mixed nodes (terminal for a code AND parent of others): {len(mixed)}")
for m in mixed[:10]: print("   ", m)
