"""On-disk store of supplementary material for a set of papers.

Supplementary material carries content a cell type report often cannot be
written without — DEG tables, cluster-to-name mappings, the marker lists name
resolution depends on. This module keeps that material on disk, one directory
per paper, alongside a manifest saying what each file and table contains
(schema: ``supplement_manifest.schema.json``).

The division of labour is deliberate. Everything here is mechanical: list what
exists, fetch bytes, unpack archives, describe the shape of a spreadsheet,
slice a region out of one. Deciding *what a table is for* needs a model to read
it, and that happens in the ``index-supplements`` skill, which drives the
functions below and writes the manifest's ``tables`` section. Nothing in this
module calls an LLM.

Two operations matter for keeping large files out of an agent's context:

``outline``
    A compact description of a file's shape — worksheet names, dimensions, the
    detected header row, a few sample rows. Bounded output regardless of how
    big the file is.

``slice``
    A targeted read of one region, so a reader that has decided which sheet it
    wants never loads the rest.

Nothing here knows about atlas-reporter's directory layout: the store root is
an argument, and paper identity comes from a CAS+ document or an explicit DOI.

.. code-block:: python

    from pathlib import Path
    from atlas_chat.services.supplement_store import (
        inventory_from_jats, adopt_manual_files, outline_file, read_slice,
    )

    files = inventory_from_jats(Path("paper.jats.xml"))
    manifest = adopt_manual_files(
        store_root=Path("supplements"),
        doi="10.1038/s41586-024-08002-x",
        incoming=Path("incoming/10.1038_s41586-024-08002-x"),
        listed=files,
    )
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import logging
import re
import shutil
import sys
import tarfile
import xml.etree.ElementTree as ET
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from atlas_chat.services.local_snippet_index import paper_slug

logger = logging.getLogger(__name__)

MANIFEST_VERSION = 1

#: Non-tabular members above this are recorded but not unpacked. Supplementary
#: videos and image stacks routinely run to hundreds of MB and never carry
#: table content.
DEFAULT_SIZE_CAP_BYTES = 80 * 1024 * 1024

#: Tabular members get their own, far higher ceiling. A DEG table for a whole
#: atlas legitimately runs to ~100 MB — in the prenatal skin atlas the largest
#: supplementary table is also the one a marker query most needs, so a general
#: size cap would exclude exactly the wrong file. ``outline`` reads workbooks
#: in openpyxl's read-only mode, so a large table costs disk, not context.
DEFAULT_TABULAR_SIZE_CAP_BYTES = 512 * 1024 * 1024

#: Media kinds that hold tables, and so get the tabular ceiling above.
TABULAR_MEDIA_TYPES = frozenset({"xlsx", "csv", "tsv", "txt", "docx"})

#: Rows shown per sheet by ``outline``. Enough to see what a table holds and to
#: spot a title row above the header, without pulling the table into context.
OUTLINE_SAMPLE_ROWS = 8

#: Columns shown per row by ``outline``. Cell-metadata tables can be very wide.
OUTLINE_MAX_COLS = 40

#: Truncation applied to every cell ``outline`` prints. Some supplement cells
#: hold whole paragraphs of methods text.
OUTLINE_CELL_CHARS = 200

#: Rows always read when looking for the header, even when the caller asked for
#: fewer sample rows. Header quality must not depend on how much the caller
#: wanted to *see*: a table whose header sits on row 2 behind two title rows
#: would otherwise be mis-detected by `--rows 2`.
HEADER_SCAN_ROWS = 6

_MEDIA_TYPES = {
    ".xlsx": "xlsx",
    ".xls": "xlsx",
    ".xlsm": "xlsx",
    ".csv": "csv",
    ".tsv": "tsv",
    ".txt": "txt",
    ".docx": "docx",
    ".doc": "docx",
    ".pdf": "pdf",
    ".zip": "zip",
    ".gz": "zip",
    ".tgz": "zip",
    ".tar": "zip",
    ".mp4": "video",
    ".mov": "video",
    ".avi": "video",
    ".png": "image",
    ".jpg": "image",
    ".jpeg": "image",
    ".tif": "image",
    ".tiff": "image",
}

_ARCHIVE_SUFFIXES = {".zip", ".tar", ".gz", ".tgz"}


class SupplementStoreError(RuntimeError):
    """Raised for unrecoverable store problems (bad paths, unreadable manifest)."""


# ------------------------------------------------------------------
# Store layout
# ------------------------------------------------------------------


def paper_dir(store_root: Path, doi: str) -> Path:
    """Directory holding one paper's supplements, named by DOI slug."""
    return Path(store_root) / "papers" / paper_slug(doi)


def manifest_path(store_root: Path, doi: str) -> Path:
    """Path to a paper's manifest JSON."""
    return paper_dir(store_root, doi) / "manifest.json"


def files_dir(store_root: Path, doi: str) -> Path:
    """Directory holding the paper's supplement bytes (git-ignored)."""
    return paper_dir(store_root, doi) / "files"


def media_type(name: str) -> str:
    """Extension-derived kind, matching the manifest's ``media_type`` vocabulary."""
    return _MEDIA_TYPES.get(Path(name).suffix.lower(), "other")


def _is_archive(name: str) -> bool:
    return Path(name).suffix.lower() in _ARCHIVE_SUFFIXES


def _now() -> str:
    return datetime.now(UTC).isoformat(timespec="seconds")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _rel(store_root: Path, path: Path) -> str:
    """Path relative to the store root, so a manifest survives being moved."""
    try:
        return str(path.resolve().relative_to(Path(store_root).resolve()))
    except ValueError:
        return str(path)


# ------------------------------------------------------------------
# Inventory: what supplements does this paper have?
# ------------------------------------------------------------------

_JATS_NS = {"xlink": "http://www.w3.org/1999/xlink"}


def _element_text(element: ET.Element | None) -> str:
    if element is None:
        return ""
    return re.sub(r"\s+", " ", "".join(element.itertext())).strip()


def inventory_from_jats(jats_path: Path) -> list[dict[str, Any]]:
    """List a paper's supplementary files from its JATS XML.

    The article XML is the cheapest source of the file list *and* the only place
    the publisher's labels and captions live. A caption like
    "Supplementary Tables 1-40." often describes the contents better than
    anything recoverable from the bytes, so it is worth reading before opening
    a single file.

    Args:
        jats_path: Path to the article's JATS XML (e.g. Europe PMC fullTextXML).

    Returns:
        One dict per supplementary file with ``file_id``, ``label``,
        ``caption``, ``media_type`` and ``status: "listed"`` — the ``files``
        entries of a manifest, before anything has been fetched.
    """
    root = ET.parse(jats_path).getroot()
    out: list[dict[str, Any]] = []
    seen: set[str] = set()

    for element in root.iter():
        if not element.tag.endswith("supplementary-material"):
            continue
        href = element.get("{http://www.w3.org/1999/xlink}href") or element.get("href")
        if not href:
            # Some publishers attach the filename to a nested <media>.
            for child in element.iter():
                if child.tag.endswith("media"):
                    href = child.get("{http://www.w3.org/1999/xlink}href") or child.get("href")
                    if href:
                        break
        if not href or href in seen:
            continue
        seen.add(href)

        label = ""
        caption = ""
        for child in element.iter():
            if child.tag.endswith("label") and not label:
                label = _element_text(child)
            elif child.tag.endswith("caption") and not caption:
                caption = _element_text(child)

        entry: dict[str, Any] = {
            "file_id": Path(href).name,
            "media_type": media_type(href),
            "status": "listed",
            "retrieval": {"route": "jats_listing"},
        }
        if label:
            entry["label"] = label
        if caption:
            entry["caption"] = caption
        out.append(entry)

    logger.info("JATS listed %d supplementary files in %s", len(out), jats_path)
    return out


# ------------------------------------------------------------------
# Getting files into the store
# ------------------------------------------------------------------


def adopt_manual_files(
    store_root: Path,
    doi: str,
    incoming: Path,
    listed: list[dict[str, Any]] | None = None,
    paper: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Take files a user dropped in by hand into the store.

    This is the only route available for closed-access papers, so a manually
    supplied file is recorded as a first-class retrieval route rather than as a
    fallback. Files are copied (not moved), so the drop directory stays
    reusable.

    Args:
        store_root: Store root; the paper's directory is created under it.
        doi: DOI of the paper the files belong to.
        incoming: Directory holding the dropped files.
        listed: Optional inventory from :func:`inventory_from_jats`, so labels
            and captions carry over to files matched by name.
        paper: Optional extra ``paper`` fields for the manifest. Only ``pmcid``
            is carried; the corpus's own metadata lives in its CAS+ document.

    Returns:
        The manifest dict, already written to disk.
    """
    incoming = Path(incoming)
    if not incoming.is_dir():
        raise SupplementStoreError(f"incoming directory does not exist: {incoming}")

    target = files_dir(store_root, doi)
    target.mkdir(parents=True, exist_ok=True)

    by_id = {entry["file_id"]: dict(entry) for entry in (listed or [])}

    for src in sorted(incoming.iterdir()):
        if not src.is_file() or src.name.startswith(".") or src.name == "README.md":
            continue
        dest = target / src.name
        if not dest.exists() or dest.stat().st_size != src.stat().st_size:
            shutil.copy2(src, dest)
        entry = by_id.setdefault(src.name, {"file_id": src.name})
        entry.update(
            {
                "media_type": media_type(src.name),
                "size_bytes": dest.stat().st_size,
                "path": _rel(store_root, dest),
                "status": "present",
                "retrieval": {
                    "route": "manual",
                    "retrieved_at": _now(),
                    "sha256": _sha256(dest),
                    "note": f"copied from {incoming}",
                },
            }
        )
        logger.info("adopted %s (%d bytes)", src.name, entry["size_bytes"])

    manifest = load_manifest(store_root, doi) or {
        "manifest_version": MANIFEST_VERSION,
        "paper": {"doi": doi},
        "files": [],
    }
    manifest["paper"].update(paper or {})
    manifest["files"] = _merge_files(manifest.get("files", []), list(by_id.values()))
    write_manifest(store_root, doi, manifest)
    return manifest


def _merge_files(
    existing: list[dict[str, Any]], incoming: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    """Merge file entries by ``file_id``; a present file never regresses to listed."""
    merged: dict[str, dict[str, Any]] = {entry["file_id"]: dict(entry) for entry in existing}
    for entry in incoming:
        current = merged.get(entry["file_id"])
        if current is None:
            merged[entry["file_id"]] = dict(entry)
            continue
        if current.get("status") == "present" and entry.get("status") == "listed":
            # Keep the bytes we have; take any label/caption the listing adds.
            for key in ("label", "caption"):
                if key in entry and key not in current:
                    current[key] = entry[key]
            continue
        current.update(entry)
    return list(merged.values())


# ------------------------------------------------------------------
# Unpacking archives
# ------------------------------------------------------------------


def unpack_archives(
    store_root: Path,
    doi: str,
    size_cap: int = DEFAULT_SIZE_CAP_BYTES,
    tabular_size_cap: int = DEFAULT_TABULAR_SIZE_CAP_BYTES,
) -> dict[str, Any]:
    """Expand archive files in the store and record their member trees.

    Nature-style supplements arrive as one zip holding dozens of workbooks, so
    the members are the interesting unit, not the archive. Oversized members and
    members that are themselves archives are listed but not expanded — a nested
    archive is recorded as a gap rather than recursed into blindly.

    Args:
        store_root: Store root.
        doi: Paper whose archives to expand.
        size_cap: Ceiling for non-tabular members (video, images).
        tabular_size_cap: Ceiling for spreadsheets and documents, which is much
            higher because a large table is usually the one worth having.

    Returns:
        The updated manifest, written to disk.
    """
    manifest = load_manifest(store_root, doi)
    if manifest is None:
        raise SupplementStoreError(f"no manifest for {doi} in {store_root}")

    # Gaps this run produces replace the previous run's gaps for the same
    # archives — otherwise raising a cap leaves the old "above the cap" gap
    # behind and the manifest reports a problem that no longer exists.
    archives = {
        entry["file_id"]
        for entry in manifest.get("files", [])
        if entry.get("status") == "present" and _is_archive(entry["file_id"])
    }
    gaps: list[dict[str, Any]] = [
        gap for gap in manifest.get("gaps", []) if gap.get("file_id") not in archives
    ]

    for entry in manifest.get("files", []):
        if entry.get("status") != "present" or not _is_archive(entry["file_id"]):
            continue
        archive = Path(store_root) / entry["path"]
        out_dir = archive.parent / f"{archive.stem}__unpacked"
        members, member_gaps = _extract(archive, out_dir, store_root, size_cap, tabular_size_cap)
        entry["members"] = members
        gaps.extend(member_gaps)
        logger.info("unpacked %s → %d members", entry["file_id"], len(members))

    manifest["gaps"] = _dedupe_gaps(gaps)
    write_manifest(store_root, doi, manifest)
    return manifest


def _extract(
    archive: Path,
    out_dir: Path,
    store_root: Path,
    size_cap: int,
    tabular_size_cap: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Extract one archive; return (member entries, gaps)."""
    members: list[dict[str, Any]] = []
    gaps: list[dict[str, Any]] = []
    out_dir.mkdir(parents=True, exist_ok=True)

    for name, size in _archive_listing(archive):
        if name.endswith("/"):
            continue
        kind = media_type(name)
        entry: dict[str, Any] = {
            "member_path": name,
            "media_type": kind,
            "size_bytes": size,
            "extracted": False,
        }
        cap = tabular_size_cap if kind in TABULAR_MEDIA_TYPES else size_cap
        cap_flag = "--tabular-size-cap" if kind in TABULAR_MEDIA_TYPES else "--size-cap"
        if size > cap:
            gaps.append(
                {
                    "file_id": archive.name,
                    "member_path": name,
                    "reason": f"member is {size} bytes, above the {cap}-byte cap; not unpacked",
                    "action": f"re-run with a higher {cap_flag} if this file is needed",
                }
            )
        elif _is_archive(name):
            gaps.append(
                {
                    "file_id": archive.name,
                    "member_path": name,
                    "reason": "member is itself an archive; nested archives are not expanded",
                    "action": "extract it by hand into incoming/ and re-run adopt",
                }
            )
        else:
            dest = _safe_extract_one(archive, name, out_dir)
            if dest is None:
                gaps.append(
                    {
                        "file_id": archive.name,
                        "member_path": name,
                        "reason": "member path escapes the archive root; refused",
                    }
                )
            else:
                entry["extracted"] = True
                entry["path"] = _rel(store_root, dest)
        members.append(entry)

    return members, gaps


def _archive_listing(archive: Path) -> list[tuple[str, int]]:
    if zipfile.is_zipfile(archive):
        with zipfile.ZipFile(archive) as zf:
            return [(info.filename, info.file_size) for info in zf.infolist()]
    if tarfile.is_tarfile(archive):
        with tarfile.open(archive) as tf:
            return [(m.name, m.size) for m in tf.getmembers() if m.isfile()]
    raise SupplementStoreError(f"not a zip or tar archive: {archive}")


def _safe_extract_one(archive: Path, name: str, out_dir: Path) -> Path | None:
    """Extract one member, refusing paths that escape ``out_dir``."""
    dest = (out_dir / name).resolve()
    if not str(dest).startswith(str(out_dir.resolve())):
        return None
    dest.parent.mkdir(parents=True, exist_ok=True)
    if zipfile.is_zipfile(archive):
        with zipfile.ZipFile(archive) as zf, zf.open(name) as src, dest.open("wb") as out:
            shutil.copyfileobj(src, out)
    else:
        with tarfile.open(archive) as tf:
            handle = tf.extractfile(name)
            if handle is None:
                return None
            with handle, dest.open("wb") as out:
                shutil.copyfileobj(handle, out)
    return dest


def _dedupe_gaps(gaps: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, str, str]] = set()
    out: list[dict[str, Any]] = []
    for gap in gaps:
        key = (gap.get("file_id", ""), gap.get("member_path", ""), gap["reason"])
        if key not in seen:
            seen.add(key)
            out.append(gap)
    return out


# ------------------------------------------------------------------
# Outline: the shape of a file, bounded
# ------------------------------------------------------------------


def outline_file(
    path: Path,
    sample_rows: int = OUTLINE_SAMPLE_ROWS,
    max_cols: int = OUTLINE_MAX_COLS,
) -> dict[str, Any]:
    """Describe a file's shape without loading it into context.

    Output size is bounded by ``sample_rows`` and ``max_cols`` however large
    the file is. For a workbook this is one block per worksheet; the reader can
    usually tell from the sheet names and the first rows which sheet is the
    legend and which hold DEG results.

    Args:
        path: File to describe.
        sample_rows: Rows to show per table.
        max_cols: Columns to show per row.

    No limit here is silent. Each table reports its true ``n_rows`` and
    ``n_cols`` alongside ``truncated_rows`` / ``truncated_cols``, so a caller
    can always tell the difference between "this is the whole table" and "this
    is the first few rows of it" — and knows to reach for :func:`read_slice`
    rather than assuming it has seen everything.

    Returns:
        ``{"path", "media_type", "tables": [...]}`` where each table carries
        ``locator``, ``n_rows``, ``n_cols``, ``truncated_rows``,
        ``truncated_cols``, ``header_row_guess`` and ``rows``. Unsupported kinds
        return an empty ``tables`` list and a ``note``.
    """
    path = Path(path)
    if not path.is_file():
        # An empty table list for a path that does not exist would read as
        # "this file holds nothing" — the absence-as-negative-result trap.
        raise SupplementStoreError(f"no such file: {path}")
    kind = media_type(path.name)
    out: dict[str, Any] = {"path": str(path), "media_type": kind, "tables": []}

    if kind == "xlsx":
        out["tables"] = _outline_xlsx(path, sample_rows, max_cols)
    elif kind in {"csv", "tsv", "txt"}:
        out["tables"] = _outline_delimited(path, kind, sample_rows, max_cols)
    elif kind == "docx":
        out["tables"] = _outline_docx(path, sample_rows, max_cols)
    else:
        out["note"] = (
            f"no outline support for {kind}; the manifest records this file without a content index"
        )
    return out


def _clip(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return text[:OUTLINE_CELL_CHARS] + "…" if len(text) > OUTLINE_CELL_CHARS else text


def _trim(row: list[str]) -> list[str]:
    """Drop trailing empty cells.

    Spreadsheet readers pad rows out to the sheet's width, and supplement
    sheets are often declared far wider than their data. Padding every sampled
    row with dozens of empty strings is noise a reader has to look past.
    """
    end = len(row)
    while end and not row[end - 1]:
        end -= 1
    return row[:end]


def _looks_numeric(cell: str) -> bool:
    try:
        float(cell.replace(",", ""))
    except ValueError:
        return False
    return True


def _header_row_guess(rows: list[list[str]], scan: int = 6) -> int:
    """Index of the most plausible header row.

    Publisher tables are rarely a bare grid. The prenatal skin supplements are
    typical: a title row ("DEG analysis for macrophage subpopulations…"), then
    sometimes a group-label row naming the block of columns beneath it, then the
    real header, then data.

    Two signals separate the header from everything above and below it. It is
    the widest row near the top — wider than a title, which occupies one cell —
    and its cells are *words*, where the data rows below are mostly numbers.
    Column count alone picks the first data row for a table whose header is no
    wider than its data, so both signals are needed.

    Falls back to the widest row when no row looks textual (a table of gene
    symbols with no header at all).
    """
    candidates: list[tuple[int, int]] = []  # (filled cells, index)
    counts: list[int] = []

    for row in rows[:scan]:
        filled = [cell for cell in row if cell]
        counts.append(len(filled))
        if len(filled) < 2:
            continue
        textual = sum(1 for cell in filled if not _looks_numeric(cell))
        if textual / len(filled) >= 0.6:
            candidates.append((len(filled), len(counts) - 1))

    if candidates:
        # Widest textual row wins; ties go to the earliest, which is the header
        # rather than a repeat of it.
        return max(candidates, key=lambda pair: (pair[0], -pair[1]))[1]

    if not counts:
        return 0
    # Nothing looked textual — a header whose column names are themselves
    # numbers (cluster ids 0..4, timepoints). Take the earliest row that is
    # within one cell of the widest: such a header is often one narrower than
    # its data rows, because the index column's own header is blank.
    threshold = max(counts) - 1
    return next(index for index, count in enumerate(counts) if count >= threshold)


def _outline_xlsx(path: Path, sample_rows: int, max_cols: int) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise SupplementStoreError(
            "reading .xlsx needs openpyxl — install the [supplements] extra"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    tables: list[dict[str, Any]] = []
    try:
        scan_rows = max(sample_rows, HEADER_SCAN_ROWS)
        for sheet in workbook.worksheets:
            scanned: list[list[str]] = []
            for row in sheet.iter_rows(max_row=scan_rows, max_col=max_cols, values_only=True):
                scanned.append(_trim([_clip(cell) for cell in row]))
            rows = scanned[:sample_rows]
            tables.append(
                {
                    "locator": sheet.title,
                    "n_rows": sheet.max_row,
                    "n_cols": sheet.max_column,
                    "truncated_cols": bool(sheet.max_column and sheet.max_column > max_cols),
                    "truncated_rows": bool(sheet.max_row and sheet.max_row > len(rows)),
                    "header_row_guess": _header_row_guess(scanned),
                    "rows": rows,
                }
            )
    finally:
        workbook.close()
    return tables


def _sniff_delimiter(sample: str, kind: str) -> str:
    if kind == "tsv":
        return "\t"
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        return "\t" if sample.count("\t") > sample.count(",") else ","


def _outline_delimited(
    path: Path, kind: str, sample_rows: int, max_cols: int
) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8", errors="replace")
    delimiter = _sniff_delimiter(text[:8192], kind)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    scan_rows = max(sample_rows, HEADER_SCAN_ROWS)
    scanned: list[list[str]] = []
    total = 0
    widest = 0
    for index, row in enumerate(reader):
        total = index + 1
        widest = max(widest, len(row))
        if index < scan_rows:
            scanned.append(_trim([_clip(cell) for cell in row[:max_cols]]))
    rows = scanned[:sample_rows]
    return [
        {
            "locator": path.name,
            "delimiter": delimiter,
            "n_rows": total,
            "n_cols": widest,
            "truncated_cols": widest > max_cols,
            "truncated_rows": total > len(rows),
            "header_row_guess": _header_row_guess(scanned),
            "rows": rows,
        }
    ]


def _outline_docx(path: Path, sample_rows: int, max_cols: int) -> list[dict[str, Any]]:
    """Tables in a .docx, read straight from the package XML.

    Supplementary Information documents carry cluster-annotation tables often
    available nowhere else. Parsing ``word/document.xml`` avoids a python-docx
    dependency for what is a shallow read.
    """
    if not zipfile.is_zipfile(path):
        return []
    with zipfile.ZipFile(path) as zf:
        try:
            xml = zf.read("word/document.xml")
        except KeyError:
            return []

    ns = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    root = ET.fromstring(xml)
    tables: list[dict[str, Any]] = []
    scan_rows = max(sample_rows, HEADER_SCAN_ROWS)
    for index, table in enumerate(root.iter(f"{ns}tbl"), start=1):
        scanned: list[list[str]] = []
        all_rows = list(table.iter(f"{ns}tr"))
        widest = 0
        for row_index, row in enumerate(all_rows):
            cells = [_element_text(cell) for cell in row.iter(f"{ns}tc")]
            widest = max(widest, len(cells))
            if row_index < scan_rows:
                scanned.append(_trim([_clip(cell) for cell in cells[:max_cols]]))
        rows = scanned[:sample_rows]
        tables.append(
            {
                "locator": f"table {index}",
                "n_rows": len(all_rows),
                "n_cols": widest,
                "truncated_cols": widest > max_cols,
                "truncated_rows": len(all_rows) > len(rows),
                "header_row_guess": _header_row_guess(scanned),
                "rows": rows,
            }
        )
    return tables


# ------------------------------------------------------------------
# Text: reading a legend document
# ------------------------------------------------------------------


def extract_text(path: Path, max_chars: int = 40_000) -> dict[str, Any]:
    """Plain text of a document, truncated to a budget.

    The reason this exists: a bundle of forty supplementary tables usually ships
    with a legends document describing every one of them, and it is prose, not a
    table — so :func:`outline_file` sees nothing in it. Reading that one file
    characterises the whole bundle, which makes it the highest-leverage read in
    the store.

    Args:
        path: A ``.docx``, ``.txt`` or other text-bearing file.
        max_chars: Truncation budget. Legend documents run to a few tens of
            thousands of characters; the default holds a typical one whole.

    Returns:
        ``{"path", "media_type", "text", "chars", "truncated"}``. Unsupported
        kinds return empty text and a ``note`` rather than raising, so a caller
        can record a gap and move on.
    """
    path = Path(path)
    if not path.is_file():
        raise SupplementStoreError(f"no such file: {path}")
    kind = media_type(path.name)
    out: dict[str, Any] = {"path": str(path), "media_type": kind}

    if kind == "docx":
        text = _docx_text(path)
    elif kind in {"txt", "csv", "tsv"}:
        text = path.read_text(encoding="utf-8", errors="replace")
    else:
        out.update(
            {
                "text": "",
                "chars": 0,
                "truncated": False,
                "note": (
                    f"no text extraction for {kind}; PDFs need the [local-index] "
                    "extra and services._pdf_parser"
                ),
            }
        )
        return out

    out["chars"] = len(text)
    out["truncated"] = len(text) > max_chars
    out["text"] = text[:max_chars]
    return out


def _docx_text(path: Path) -> str:
    """Paragraph text of a .docx, one paragraph per line."""
    if not zipfile.is_zipfile(path):
        return ""
    with zipfile.ZipFile(path) as zf:
        try:
            xml = zf.read("word/document.xml")
        except KeyError:
            return ""

    ns = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    root = ET.fromstring(xml)
    lines = [_element_text(para) for para in root.iter(f"{ns}p")]
    return "\n".join(line for line in lines if line)


# ------------------------------------------------------------------
# Slice: a targeted read
# ------------------------------------------------------------------


def read_slice(
    path: Path,
    locator: str | None = None,
    start: int = 0,
    limit: int = 50,
    columns: list[str] | None = None,
    header_row: int | None = None,
) -> dict[str, Any]:
    """Read one region of a table.

    The counterpart to :func:`outline_file`: once a reader knows which sheet it
    wants, this pulls just that part. Row indices are 0-based over the file's
    rows, header included, so they line up with ``outline``'s
    ``header_row_guess``.

    Args:
        path: File to read.
        locator: Worksheet name (xlsx) or table locator (docx). Ignored for
            single-table files; defaults to the first table.
        start: First row to return, 0-based, counted from the file's first row.
        limit: Maximum rows to return.
        columns: Restrict to these column names. Requires ``header_row`` (or
            relies on the guess) so headers can be resolved.
        header_row: 0-based header row index. Defaults to the guess.

    Returns:
        ``{"path", "locator", "header", "rows", "start", "returned", "n_rows"}``.
    """
    path = Path(path)
    outline = outline_file(path, sample_rows=max(limit + start, OUTLINE_SAMPLE_ROWS))
    tables = outline["tables"]
    if not tables:
        raise SupplementStoreError(f"nothing tabular to slice in {path}")

    table = _pick_table(tables, locator)
    effective_header = table["header_row_guess"] if header_row is None else header_row
    rows = _read_rows(path, table, start, limit)
    header = _read_header(path, table, effective_header)

    if columns:
        keep = [header.index(name) for name in columns if name in header]
        missing = [name for name in columns if name not in header]
        if missing:
            raise SupplementStoreError(
                f"columns not in header {header!r}: {missing!r}",
            )
        header = [header[i] for i in keep]
        rows = [[row[i] if i < len(row) else "" for i in keep] for row in rows]

    return {
        "path": str(path),
        "locator": table["locator"],
        "header_row": effective_header,
        "header": header,
        "n_rows": table["n_rows"],
        "start": start,
        "returned": len(rows),
        "rows": rows,
    }


def _pick_table(tables: list[dict[str, Any]], locator: str | None) -> dict[str, Any]:
    if locator is None:
        return tables[0]
    for table in tables:
        if table["locator"] == locator:
            return table
    available = [table["locator"] for table in tables]
    raise SupplementStoreError(f"no table {locator!r}; available: {available!r}")


def _read_rows(path: Path, table: dict[str, Any], start: int, limit: int) -> list[list[str]]:
    kind = media_type(path.name)
    if kind == "xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[table["locator"]]
            rows: list[list[str]] = []
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index < start:
                    continue
                if len(rows) >= limit:
                    break
                rows.append(_trim([_clip(cell) for cell in row]))
            return rows
        finally:
            workbook.close()

    # Delimited and docx files are re-outlined with a large sample; cheap
    # relative to the size of file that reaches this path.
    full = outline_file(path, sample_rows=start + limit, max_cols=10_000)
    picked = _pick_table(full["tables"], table["locator"])
    return picked["rows"][start : start + limit]


def _read_header(path: Path, table: dict[str, Any], header_row: int) -> list[str]:
    rows = _read_rows(path, table, header_row, 1)
    return rows[0] if rows else []


# ------------------------------------------------------------------
# Manifest IO
# ------------------------------------------------------------------


def load_manifest(store_root: Path, doi: str) -> dict[str, Any] | None:
    """Read a paper's manifest, or None if the store has no entry for it."""
    path = manifest_path(store_root, doi)
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise SupplementStoreError(f"manifest is not valid JSON: {path}") from exc


def write_manifest(store_root: Path, doi: str, manifest: dict[str, Any]) -> Path:
    """Validate a manifest against its schema and write it.

    Raises:
        SupplementStoreError: If the manifest does not conform. Writing an
            invalid manifest is worse than failing — downstream steps trust it.
    """
    validate_manifest(manifest)
    path = manifest_path(store_root, doi)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    logger.info("wrote %s", path)
    return path


def validate_manifest(manifest: dict[str, Any]) -> None:
    """Validate against ``supplement_manifest.schema.json``.

    Raises:
        SupplementStoreError: On the first schema violation, with the JSON path.
    """
    import jsonschema  # type: ignore[import-untyped]

    from atlas_chat.schemas import load_schema

    try:
        jsonschema.validate(manifest, load_schema("supplement_manifest.schema.json"))
    except jsonschema.ValidationError as exc:
        location = "/".join(str(part) for part in exc.absolute_path) or "(root)"
        raise SupplementStoreError(f"manifest invalid at {location}: {exc.message}") from exc


def cross_check_manifest(manifest: dict[str, Any]) -> list[str]:
    """Consistency rules the schema cannot express.

    The schema pins the shape; these are the rules that make a manifest
    *usable* — a table pointing at a file that isn't there, or a present file
    with no path, both pass schema validation and fail a reader.

    Returns:
        Human-readable problems, empty if the manifest is consistent.
    """
    problems: list[str] = []
    files = {entry["file_id"]: entry for entry in manifest.get("files", [])}

    for entry in manifest.get("files", []):
        if entry["status"] == "present" and not entry.get("path"):
            problems.append(f"file {entry['file_id']} is 'present' but has no path")
        if entry["status"] != "present" and entry.get("path"):
            problems.append(f"file {entry['file_id']} is '{entry['status']}' but has a path")

    for index, table in enumerate(manifest.get("tables", [])):
        parent = files.get(table["file_id"])
        if parent is None:
            problems.append(f"tables[{index}] points at unknown file_id {table['file_id']!r}")
            continue
        if parent["status"] != "present":
            problems.append(
                f"tables[{index}] points at file {table['file_id']} "
                f"with status '{parent['status']}'"
            )
        member = table.get("member_path")
        if member:
            known = {m["member_path"] for m in parent.get("members", [])}
            if known and member not in known:
                problems.append(
                    f"tables[{index}] member_path {member!r} is not a member of {table['file_id']}"
                )
        elif parent.get("media_type") == "zip":
            problems.append(
                f"tables[{index}] is inside archive {table['file_id']} but names no member_path"
            )
    return problems


def touch_indexed_at(store_root: Path, doi: str) -> dict[str, Any]:
    """Stamp ``indexed_at`` after the content index has been (re)written."""
    manifest = load_manifest(store_root, doi)
    if manifest is None:
        raise SupplementStoreError(f"no manifest for {doi} in {store_root}")
    manifest["indexed_at"] = _now()
    write_manifest(store_root, doi, manifest)
    return manifest


# ------------------------------------------------------------------
# CAS+ helpers
# ------------------------------------------------------------------


def corpus_papers(cas: dict[str, Any]) -> list[dict[str, Any]]:
    """The papers whose supplements a project needs, from a CAS+ document.

    The atlas paper plus every subatlas paper that has a DOI. Takes the parsed
    document, not a project name, so it works for any CAS+ source.

    Returns:
        Dicts with ``doi``, ``role`` and (where known) ``title``, ``pmcid``.
        This describes the corpus, not a manifest: ``role`` tells a caller which
        papers to fetch for, and is deliberately not written into the store.
    """
    source = cas.get("source", {})
    papers: list[dict[str, Any]] = []
    if source.get("doi"):
        entry = {"doi": source["doi"], "role": "atlas"}
        for key in ("title", "pmcid", "pmid"):
            if source.get(key):
                entry[key] = source[key]
        papers.append(entry)
    for sub in source.get("subatlas_papers", []):
        doi = sub.get("doi")
        if doi:
            papers.append({"doi": doi, "role": "subatlas", "title": sub.get("label", "")})
    return papers


# ------------------------------------------------------------------
# CLI
# ------------------------------------------------------------------


def _print(payload: Any) -> None:
    print(json.dumps(payload, indent=2))


def _cmd_inventory(args: argparse.Namespace) -> int:
    _print(inventory_from_jats(Path(args.jats)))
    return 0


def _cmd_adopt(args: argparse.Namespace) -> int:
    listed = None
    if args.jats:
        listed = inventory_from_jats(Path(args.jats))
    paper: dict[str, Any] = {}
    if args.pmcid:
        paper["pmcid"] = args.pmcid
    manifest = adopt_manual_files(
        store_root=Path(args.store),
        doi=args.doi,
        incoming=Path(args.incoming),
        listed=listed,
        paper=paper,
    )
    _print(
        {
            "manifest": str(manifest_path(Path(args.store), args.doi)),
            "files": len(manifest["files"]),
        }
    )
    return 0


def _cmd_unpack(args: argparse.Namespace) -> int:
    manifest = unpack_archives(
        Path(args.store),
        args.doi,
        size_cap=args.size_cap,
        tabular_size_cap=args.tabular_size_cap,
    )
    _print(
        {
            "files": [
                {"file_id": f["file_id"], "members": len(f.get("members", []))}
                for f in manifest["files"]
            ],
            "gaps": manifest.get("gaps", []),
        }
    )
    return 0


def _cmd_outline(args: argparse.Namespace) -> int:
    _print(outline_file(Path(args.file), sample_rows=args.rows, max_cols=args.cols))
    return 0


def _cmd_text(args: argparse.Namespace) -> int:
    _print(extract_text(Path(args.file), max_chars=args.max_chars))
    return 0


def _cmd_slice(args: argparse.Namespace) -> int:
    _print(
        read_slice(
            Path(args.file),
            locator=args.locator,
            start=args.start,
            limit=args.limit,
            columns=args.columns,
            header_row=args.header_row,
        )
    )
    return 0


def _cmd_show(args: argparse.Namespace) -> int:
    manifest = load_manifest(Path(args.store), args.doi)
    if manifest is None:
        print(f"no manifest for {args.doi} in {args.store}", file=sys.stderr)
        return 1
    _print(manifest)
    return 0


def _cmd_check(args: argparse.Namespace) -> int:
    manifest = load_manifest(Path(args.store), args.doi)
    if manifest is None:
        print(f"no manifest for {args.doi} in {args.store}", file=sys.stderr)
        return 1
    try:
        validate_manifest(manifest)
    except SupplementStoreError as exc:
        print(str(exc), file=sys.stderr)
        return 1
    problems = cross_check_manifest(manifest)
    if problems:
        for problem in problems:
            print(problem, file=sys.stderr)
        return 1
    print("manifest OK")
    return 0


def _cmd_papers(args: argparse.Namespace) -> int:
    cas = json.loads(Path(args.cas).read_text(encoding="utf-8"))
    _print(corpus_papers(cas))
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="atlas_chat.cli_supplements",
        description=(
            "Supplement store: list, adopt, unpack, outline and slice a paper's "
            "supplementary material. Deciding what a table is FOR is the "
            "index-supplements skill's job; this is the mechanical half."
        ),
    )
    sub = parser.add_subparsers(dest="command", required=True)

    def add_store(p: argparse.ArgumentParser) -> None:
        p.add_argument("--store", required=True, help="store root directory")
        p.add_argument("--doi", required=True, help="DOI of the paper")

    inv = sub.add_parser("inventory", help="list supplementary files from a JATS XML")
    inv.add_argument("--jats", required=True)
    inv.set_defaults(func=_cmd_inventory)

    adopt = sub.add_parser("adopt", help="take manually-dropped files into the store")
    add_store(adopt)
    adopt.add_argument("--incoming", required=True, help="directory holding the dropped files")
    adopt.add_argument("--jats", help="JATS XML, to carry over labels and captions")
    adopt.add_argument("--pmcid", help="cached for the fetch routes")
    adopt.set_defaults(func=_cmd_adopt)

    unpack = sub.add_parser("unpack", help="expand archives and record member trees")
    add_store(unpack)
    unpack.add_argument(
        "--size-cap",
        type=int,
        default=DEFAULT_SIZE_CAP_BYTES,
        help="ceiling for non-tabular members (video, images)",
    )
    unpack.add_argument(
        "--tabular-size-cap",
        type=int,
        default=DEFAULT_TABULAR_SIZE_CAP_BYTES,
        help="ceiling for spreadsheets and documents",
    )
    unpack.set_defaults(func=_cmd_unpack)

    outline = sub.add_parser("outline", help="bounded description of a file's shape")
    outline.add_argument("--file", required=True)
    outline.add_argument("--rows", type=int, default=OUTLINE_SAMPLE_ROWS)
    outline.add_argument("--cols", type=int, default=OUTLINE_MAX_COLS)
    outline.set_defaults(func=_cmd_outline)

    text = sub.add_parser("text", help="plain text of a document (e.g. a table-legends .docx)")
    text.add_argument("--file", required=True)
    text.add_argument("--max-chars", type=int, default=40_000)
    text.set_defaults(func=_cmd_text)

    sl = sub.add_parser("slice", help="read one region of one table")
    sl.add_argument("--file", required=True)
    sl.add_argument("--locator", help="worksheet name or table locator")
    sl.add_argument("--start", type=int, default=0)
    sl.add_argument("--limit", type=int, default=50)
    sl.add_argument("--columns", nargs="*")
    sl.add_argument("--header-row", type=int)
    sl.set_defaults(func=_cmd_slice)

    show = sub.add_parser("show", help="print a paper's manifest")
    add_store(show)
    show.set_defaults(func=_cmd_show)

    check = sub.add_parser("check", help="validate a manifest against its schema")
    add_store(check)
    check.set_defaults(func=_cmd_check)

    papers = sub.add_parser("papers", help="corpus papers from a CAS+ document")
    papers.add_argument("--cas", required=True)
    papers.set_defaults(func=_cmd_papers)

    return parser


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    args = build_parser().parse_args(argv)
    try:
        return int(args.func(args))
    except SupplementStoreError as exc:
        print(str(exc), file=sys.stderr)
        return 1


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
